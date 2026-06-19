"""Prepara el dataset de madurez Hass (Mendeley DOI 10.17632/3xd9n945v8.1) para entrenar.

El dataset NO viene en carpetas por clase: son imágenes sueltas + una planilla (Excel/CSV) que
asocia cada foto a su etapa del Índice de Maduración (1-5). torchvision.datasets.ImageFolder
—que usa scripts/train_vision.py— exige carpetas = clases. Este script lee la planilla y copia
(o enlaza) cada imagen a data/vision/madurez/<clave_canónica>/ según el mapeo verificado:

    1 Underripe           -> madurez_verde
    2 Breaking            -> madurez_pinton
    3 Ripe (First Stage)  -> madurez_maduro_inicial
    4 Ripe (Second Stage) -> madurez_maduro_optimo
    5 Overripe            -> madurez_sobremaduro

Licencia del dataset: CC BY 4.0 (uso comercial con atribución). Atribución en docs/VISION.md.

Uso:
    # 1) descarga el dataset de https://data.mendeley.com/datasets/3xd9n945v8/1 a una carpeta
    # 2) python scripts/prepare_maturity_dataset.py --src data/vision/_raw_madurez --out data/vision/madurez
    # 3) uv run python scripts/train_vision.py --data-dir data/vision/madurez ...

NOTA: syntax-checked pero NO probado contra el archivo real (aún no descargado). Si los nombres de
columnas difieren, el script imprime las columnas detectadas; ajusta FILENAME_HINTS / STAGE_HINTS.
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import sys
from pathlib import Path

# Consola Windows en cp1252 no encodea ✓/emojis → fuerza UTF-8 (evita crash en los prints).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Mapeo verificado etapa del dataset -> clave canónica de src/avorag/vision/labels.py
STAGE_TO_KEY: dict[str, str] = {
    "1": "madurez_verde",
    "underripe": "madurez_verde",
    "under-ripe": "madurez_verde",
    "2": "madurez_pinton",
    "breaking": "madurez_pinton",
    "3": "madurez_maduro_inicial",
    "ripe first stage": "madurez_maduro_inicial",
    "ripe (first stage)": "madurez_maduro_inicial",
    "first stage": "madurez_maduro_inicial",
    "4": "madurez_maduro_optimo",
    "ripe second stage": "madurez_maduro_optimo",
    "ripe (second stage)": "madurez_maduro_optimo",
    "second stage": "madurez_maduro_optimo",
    "5": "madurez_sobremaduro",
    "overripe": "madurez_sobremaduro",
    "over-ripe": "madurez_sobremaduro",
}
CANON_KEYS = (
    "madurez_verde",
    "madurez_pinton",
    "madurez_maduro_inicial",
    "madurez_maduro_optimo",
    "madurez_sobremaduro",
)
FILENAME_HINTS = ("file", "image", "photo", "name", "archivo", "imagen", "foto")
STAGE_HINTS = ("class", "stage", "ripen", "index", "madur", "etapa", "clasif")
IMG_EXTS = (".jpg", ".jpeg", ".png", ".bmp", ".webp")


def _norm(s: object) -> str:
    return str(s).strip().lower()


def _read_rows(label_file: Path) -> list[dict[str, str]]:
    """Lee la planilla (CSV con stdlib; XLSX con openpyxl si está instalado)."""
    if label_file.suffix.lower() == ".csv":
        with label_file.open(encoding="utf-8-sig", newline="") as f:
            return [dict(r) for r in csv.DictReader(f)]
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            f"Para leer {label_file.name} (Excel) instala openpyxl (`uv pip install openpyxl`),\n"
            "o exporta la planilla a .csv y vuelve a correr este script."
        )
    wb = load_workbook(label_file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [_norm(c) if c is not None else "" for c in next(rows_iter)]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        out.append(
            {header[i]: ("" if v is None else str(v)) for i, v in enumerate(row) if i < len(header)}
        )
    return out


def _pick_column(headers: list[str], hints: tuple[str, ...]) -> str | None:
    for h in headers:
        if any(hint in h for hint in hints):
            return h
    return None


def main() -> None:
    ap = argparse.ArgumentParser(description="Organiza el dataset de madurez Hass en ImageFolder.")
    ap.add_argument(
        "--src",
        type=Path,
        required=True,
        help="Carpeta del dataset descargado (imágenes + planilla).",
    )
    ap.add_argument("--out", type=Path, default=Path("data/vision/madurez"))
    ap.add_argument(
        "--link", action="store_true", help="Enlaza (hardlink) en vez de copiar (ahorra disco)."
    )
    ap.add_argument("--dry-run", action="store_true", help="Solo cuenta; no escribe.")
    args = ap.parse_args()

    if not args.src.exists():
        sys.exit(f"No existe --src {args.src}")

    # 1) localizar la planilla de etiquetas
    sheets = sorted(p for p in args.src.rglob("*") if p.suffix.lower() in (".csv", ".xlsx", ".xls"))
    if not sheets:
        sys.exit(
            f"No encontré planilla (.csv/.xlsx) en {args.src}. El dataset trae una con las etapas."
        )
    label_file = sheets[0]
    print(f"Planilla: {label_file}")
    rows = _read_rows(label_file)
    if not rows:
        sys.exit("La planilla está vacía.")
    headers = list(rows[0].keys())
    print(f"Columnas detectadas: {headers}")

    fn_col = _pick_column(headers, FILENAME_HINTS)
    st_col = _pick_column(headers, STAGE_HINTS)
    if not fn_col or not st_col:
        sys.exit(
            f"No pude identificar columnas de archivo/etapa (archivo={fn_col}, etapa={st_col}).\n"
            "Edita FILENAME_HINTS / STAGE_HINTS en este script según las columnas de arriba."
        )
    print(f"Columna de archivo: {fn_col!r} | columna de etapa: {st_col!r}")

    # 2) índice de imágenes en disco por nombre (la planilla puede no traer ruta completa)
    by_name: dict[str, Path] = {}
    for p in args.src.rglob("*"):
        if p.suffix.lower() in IMG_EXTS:
            by_name.setdefault(p.name.lower(), p)

    counts: dict[str, int] = {}
    missing = unmapped = 0
    for r in rows:
        raw_name = (r.get(fn_col) or "").strip()
        key = STAGE_TO_KEY.get(_norm(r.get(st_col) or ""))
        if not raw_name or key is None:
            unmapped += 1
            continue
        name = Path(raw_name).name.lower()
        if name not in by_name and not name.endswith(IMG_EXTS):
            # la planilla puede omitir la extensión
            name = next((f"{name}{e}" for e in IMG_EXTS if f"{name}{e}" in by_name), name)
        src_img = by_name.get(name)
        if src_img is None:
            missing += 1
            continue
        counts[key] = counts.get(key, 0) + 1
        if args.dry_run:
            continue
        dst_dir = args.out / key
        dst_dir.mkdir(parents=True, exist_ok=True)
        dst = dst_dir / src_img.name
        if dst.exists():
            continue
        if args.link:
            try:
                os.link(src_img, dst)
            except OSError:
                shutil.copy2(src_img, dst)
        else:
            shutil.copy2(src_img, dst)

    print("\nResumen por clase:")
    for key in CANON_KEYS:
        print(f"  {key:24s} {counts.get(key, 0)}")
    print(
        f"Filas sin mapear (etapa/archivo): {unmapped} | imágenes referidas no encontradas: {missing}"
    )
    if args.dry_run:
        print(f"\n(dry-run: no se escribió nada). Quita --dry-run para materializar en {args.out}")
    else:
        print(
            f"\n✓ Listo en {args.out}. Ahora: uv run python scripts/train_vision.py --data-dir {args.out}"
        )


if __name__ == "__main__":
    main()
